"""
Created on Mon May 20 19:44:03 2019

@author: achauhan
"""

import math
from bs4 import BeautifulSoup
import urllib
import mysql.connector
from openpyxl import load_workbook
import pandas as pd
import csv
import holidays
import numpy as np
import datetime as dt
import calendar
from datetime import date, timedelta
import smtplib


# This function append entries in excel file namely Exposure_3
def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')
    
    
    try:
        writer.book = load_workbook(filename)
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            idx = writer.book.sheetnames.index(sheet_name)
            writer.book.remove(writer.book.worksheets[idx])
            writer.book.create_sheet(sheet_name, idx)
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        pass

    if startrow is None:
        startrow = 0
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)
    writer.save()
      
# This function calculate all working days between any two dates, if we have index holiday for any date then we need to add holiday(HOLIDAYS_US.append({'30-03-2018':'Good Friday'})) and run the code
    
def workdays(d, end, excluded=(6, 7)):
    days = []
    HOLIDAYS_US = holidays.US(years=[2017,2018,2019,2020])
    HOLIDAYS_US.append({'30-03-2018':'Good Friday'})
    HOLIDAYS_US.append({'19-04-2019':'Good Friday'}) 
    while d.date() <= end.date():
        if d.isoweekday() not in excluded and d not in HOLIDAYS_US:
            days.append(d)
        d += dt.timedelta(days=1)
    return days

# need to add holiday at two places
Start_Date=dt.date(year = 2019, month = 7, day = 9)
ONE_DAY = dt.timedelta(days=1)
HOLIDAYS_US = holidays.US(years=[2017,2018,2019,2020])
HOLIDAYS_US.append({'30-03-2018':'Good Friday'})
HOLIDAYS_US.append({'19-04-2019':'Good Friday'}) 
u=[*HOLIDAYS_US]


date_tm1=dt.datetime.today()-dt.timedelta(days=1)
while date_tm1.weekday() in holidays.WEEKEND or date_tm1 in HOLIDAYS_US:
   date_tm1 -= dt.timedelta(days=1)

RUN_MODE = "daily" #if daily then it runs for everyday else custom run 

if RUN_MODE == 'daily':
    Fact_Dates=workdays(date_tm1 , date_tm1)
else:
    CALC_START_DATE =  dt.datetime(2020,2,24)
    CALC_END_DATE =  dt.datetime(2020,2,24)# dt.datetime(YYYY,month,day) ex:month:7 for July date = 8
    Fact_Dates=workdays(CALC_START_DATE , CALC_END_DATE)
    


Fact_Dates_1=[]
for p in Fact_Dates:
    Fact_Dates_1.append(p.date())

# to send the mail
def send_mail(body):
    try:
        smtpObj = smtplib.SMTP('smtp-mail.outlook.com', 587)
    except Exception as e:
        print(e)
        smtpObj = smtplib.SMTP_SSL('smtp-mail.outlook.com', 465)
    #type(smtpObj) 
    smtpObj.ehlo()
    smtpObj.starttls()
    smtpObj.login('notifications@indxx.com', "Traffic@1234") 
    smtpObj.sendmail('notifications@indxx.com', ['skaushal@indxx.com','amalik@indxx.com','pavank@indxx.com','kartikya@indxx.com'], body) # Or recipient@outlook

    smtpObj.quit()
    pass



#we are fetching advance issues and decline issues from https://www.nasdaq.com/markets/most-active.aspx?exchange=NYSE
try:
    from selenium import webdriver 
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.common.exceptions import TimeoutException
   
    # example option: add 'incognito' command line arg to options
    option = webdriver.ChromeOptions()
    #option.add_argument("--incognito")
   
    # create new instance of chrome in incognito mode
    browser = webdriver.Chrome(executable_path='C:\\TUTTLE_AOL\\chromedriver.exe', chrome_options=option)
   
    # go to website
    browser.get("https://www.wsj.com/market-data/stocks/marketsdiary")
   
    # wait up to 10 seconds for page to load
    timeout = 15
   
    titles_element = browser.find_elements_by_xpath("//td[@class='WSJTables--table__cell--2dzGiO7q WSJTables--is-first--2Jt1dPu7 WSJTheme--table__cell--1At-VGNg ']")
    titles = [x.text for x in titles_element]
   
    print('titles:')
    print(titles)
   
   
    values_element = browser.find_elements_by_xpath("//td[@class='WSJTables--table__cell--2dzGiO7q WSJTheme--table__cell--1At-VGNg ']")
    values = [x.text for x in values_element]  # same concept as for-loop/list-comprehension above
    print('values:')
    print(values, '\n')
    browser.close()
   
    
    p=int(values[3].replace(',',''))
    q=int(values[6].replace(',',''))
except:
    body = 'Subject: Tuttle Response File'+'\n' + '\nHello, \n\nThere is some issue with Advance and Decline issues.\n' + '\nHave a nice day!'
    send_mail(body)
df_1=pd.read_csv("C:\\TUTTLE_AOL\\NYSE_decln.csv",names=['DATE','Value'])

a=dt.date.today()-dt.timedelta(days=1)
while a.weekday() in holidays.WEEKEND or a in HOLIDAYS_US:
   a -= dt.timedelta(days=1)
b=int(a.strftime('%Y%m%d'))



if getattr(df_1, 'DATE').isin([b]).any()==False:
    fields=[int(a.strftime('%Y%m%d')),q]
    if a in Fact_Dates_1:
        with open('C:\\TUTTLE_AOL\\NYSE_decln.csv', 'a',newline='') as f:
            writer = csv.writer(f)
            writer.writerow(fields)
                
                
    fields=[int(a.strftime('%Y%m%d')),p]       
    if a in Fact_Dates_1:
        with open('C:\\TUTTLE_AOL\\NYSE_advn.csv', 'a',newline='') as f:
            writer = csv.writer(f)
            writer.writerow(fields)
            
            
            
df_1=  pd.read_csv("C:\\TUTTLE_AOL\\NYSE_decln.csv",names=['DATE','Value'])
df_2=df_1.loc[df_1['DATE']==b]
a=df_2['Value'].values[0]
if math.isnan(a)==True or a == 0:
    body = 'Subject: Tuttle Response File'+'\n' + '\nHello, \n\n Advance and Decline issues are not available\n' + '\nHave a nice day!'
    send_mail(body)       
else:
    # calculation of Exposure
    try:
        for Calc_Date in Fact_Dates_1:
            mydb = mysql.connector.connect(
                    host="146.20.65.208",
                    user="admin5",
                    passwd="Admin@1234",
                    database="admin_icai5"
                    )   
        
            mycursor = mydb.cursor()
            
            mycursor.execute('''SELECT * FROM admin_icai5.tbl_indxx_value a WHERE a.INDXX_ID IN (554,555,556,557,558) and a.date <=%s order by date desc ''',(Calc_Date,))
        
            myresult = mycursor.fetchall()
            column=column = [d[0] for d in mycursor.description]
            df = pd.DataFrame(myresult,columns=column)
            df=df[[d not in u for d in df['date']]]
            TLT_DATA=df.query("code=='TLT' ")
            a=list(TLT_DATA['indxx_value'])
            c=[float(d) for d in a]
            P=[l for l in reversed(c)]
        
            XLU_DATA=df.query("code=='XLU' ")
            a=list(XLU_DATA['indxx_value'])
            c=[float(d) for d in a]
            Q=[l for l in reversed(c)]
            
            SPY_DATA=df.query("code=='SPY' ")
            a=list(SPY_DATA['indxx_value'])
            c=[float(d) for d in a]
            S=[l for l in reversed(c)]
            SPY_AVG=sum(S[-6:])/6
            PRICE_OSC=S[-1]-SPY_AVG
            
            QQQ_DATA=df.query("code=='QQQ' ")
            a=list(QQQ_DATA['indxx_value'])
            c=[float(d) for d in a]
            T=[l for l in reversed(c)]
            
            IWM_DATA=df.query("code=='IWM' ")
            a=list(IWM_DATA['indxx_value'])
            c=[float(d) for d in a]
            U=[l for l in reversed(c)]
            
            
            
            
            ONE_DAY = dt.timedelta(days=1)
            Calc_Date_P1 = Calc_Date + ONE_DAY
            while Calc_Date_P1.weekday() in holidays.WEEKEND or Calc_Date_P1 in HOLIDAYS_US:
                Calc_Date_P1 += ONE_DAY
                
            
            
           
            
            df_1=pd.read_csv("C:\\TUTTLE_AOL\\NYSE_decln.csv",names=['DATE','Value'])
            df_1['DATE'] = df_1['DATE'].astype(str)
            df_1['DATE']=pd.to_datetime(df_1['DATE'],format='%Y%m%d')
            df_2=pd.read_csv("C:\\TUTTLE_AOL\\NYSE_advn.csv",names=['DATE','Value1'])
            df_2['DATE'] = df_2['DATE'].astype(str)
            df_2['DATE']=pd.to_datetime(df_2['DATE'],format='%Y%m%d')
            df_3=pd.merge(df_1,df_2[['DATE','Value1']],on='DATE')
            df_3['AI_DI'] = df_3['Value1'].sub(df_3['Value'], axis = 0)
            
            df_3=df_3.loc[df_3['DATE'].dt.date <= Calc_Date]
            df_3.to_csv ("C:\\TUTTLE_AOL\\BREATH.csv", index = None, header=True) 
            
            BREATH_DATA=pd.read_csv(filepath_or_buffer = "C:\\TUTTLE_AOL\\BREATH.csv", encoding = "utf-8")
            df6=BREATH_DATA[len(BREATH_DATA)-28:len(BREATH_DATA)]
            R=list(df6["AI_DI"])
            BREADTH=sum(R[-10:])/10
            
            
            TLT_AV=sum(P[-66:])/66
            SPY_AV=sum(S[-66:])/66
            QQQ_AV=sum(T[-66:])/66
            IWM_AV=sum(U[-66:])/66
                    
            SH_1=(np.array([P[-66:]])-TLT_AV)**2
            SH_1_SUM=SH_1.sum(axis=1)
            SH_2=(np.array([S[-66:]])-SPY_AV)**2
            SH_2_SUM=SH_2.sum(axis=1)
            SH_3=(np.array([T[-66:]])-QQQ_AV)**2
            SH_3_SUM=SH_3.sum(axis=1)
            SH_4=(np.array([U[-66:]])-IWM_AV)**2
            SH_4_SUM=SH_4.sum(axis=1)    
                    
            X_BREADTH=sum(R[-28:])/28
            X_PRICE_OSC=S[-1]-(sum(S[-6:])/6)  
            Day_no =  dt.datetime.strptime(str(Calc_Date),'%Y-%m-%d').weekday()
            Day = calendar.day_name[Day_no]
                
            Strategy_3_rebal = Calc_Date - timedelta(days=Calc_Date.weekday())
            Strategy_3_rebal_P1=Strategy_3_rebal+ONE_DAY
            while Strategy_3_rebal_P1.weekday() in holidays.WEEKEND or Strategy_3_rebal_P1 in HOLIDAYS_US:
                Strategy_3_rebal_P1 += ONE_DAY
                
                
            if Strategy_3_rebal in HOLIDAYS_US:
                Strategy_3_rebal=Strategy_3_rebal_P1
                
            if Calc_Date==Start_Date:
                Strategy_3_rebal=Calc_Date
                
            if Calc_Date==Strategy_3_rebal:
                if (np.sqrt(SH_1_SUM/66)).item()==0:
                    SHARPE_1=0
                else:
                    SHARPE_1=(P[-1]-P[-66])/(np.sqrt(SH_1_SUM/66)).item()
                if (np.sqrt(SH_2_SUM/66)).item()==0:
                    SHARPE_2=0
                else:
                    SHARPE_2=(S[-1]-S[-66])/(np.sqrt(SH_2_SUM/66)).item()
                if (np.sqrt(SH_3_SUM/66)).item()==0:
                    SHARPE_3=0
                else:
                    SHARPE_3=(T[-1]-T[-66])/(np.sqrt(SH_3_SUM/66)).item()
                if (np.sqrt(SH_4_SUM/66)).item()==0:
                    SHARPE_4=0
                else:
                    SHARPE_4=(U[-1]-U[-66])/(np.sqrt(SH_4_SUM/66)).item()
            else:
                SHARPE_1=None
                SHARPE_2=None
                SHARPE_3=None
                SHARPE_4=None
                    
            if Calc_Date==Strategy_3_rebal:
                if (SHARPE_1>SHARPE_2 and   SHARPE_1>SHARPE_3)  or (SHARPE_1>SHARPE_2 and   SHARPE_1>SHARPE_4) or (SHARPE_1>SHARPE_3 and   SHARPE_1>SHARPE_4):
                    SIGNAL_3_LONG=1
                else:
                    SIGNAL_3_LONG=0
                if (SHARPE_2>SHARPE_1 and   SHARPE_2>SHARPE_3)  or (SHARPE_2>SHARPE_1 and   SHARPE_2>SHARPE_4) or (SHARPE_2>SHARPE_3 and   SHARPE_2>SHARPE_4):
                    SIGNAL_4_LONG=1
                else:
                    SIGNAL_4_LONG=0
                if (SHARPE_3>SHARPE_1 and   SHARPE_3>SHARPE_2)  or (SHARPE_3>SHARPE_1 and   SHARPE_3>SHARPE_4) or (SHARPE_3>SHARPE_2 and   SHARPE_3>SHARPE_4):
                    SIGNAL_5_LONG=1
                else:
                    SIGNAL_5_LONG=0
                if (SHARPE_4>SHARPE_1 and   SHARPE_4>SHARPE_2)  or (SHARPE_4>SHARPE_1 and   SHARPE_4>SHARPE_3) or (SHARPE_4>SHARPE_2 and   SHARPE_4>SHARPE_3):
                    SIGNAL_6_LONG=1
                else:
                    SIGNAL_6_LONG=0
            else:
                SIGNAL_3_LONG=0
                SIGNAL_4_LONG=0
                SIGNAL_5_LONG=0
                SIGNAL_6_LONG=0
                
                
                
            if Calc_Date==Start_Date:
                IL=1000
                SIGNAL=1
                
                Strategy_3_rebal=Calc_Date
                Index_rebal=Calc_Date
                Strategy_1_rebal=True
                Strategy_2_rebal=True
            else:
                REBAL_DATA=pd.read_excel("C:\\TUTTLE_AOL\\Exposure_3.xlsx",sheet_name='Measures')
                REBAL_DATA['Effective_Date']=pd.to_datetime(REBAL_DATA['Effective_Date'])
                
                REBAL_DATA=REBAL_DATA.loc[REBAL_DATA['Effective_Date'].dt.date == Calc_Date]
                E1_TR=REBAL_DATA['E1'].loc[REBAL_DATA['E1'].last_valid_index()]
                E2_TR=REBAL_DATA['E2'].loc[REBAL_DATA['E2'].last_valid_index()]
                E3_TR=REBAL_DATA['E3'].loc[REBAL_DATA['E3'].last_valid_index()]
                E4_TR=REBAL_DATA['E4'].loc[REBAL_DATA['E4'].last_valid_index()]
                E5_TR=REBAL_DATA['E5'].loc[REBAL_DATA['E5'].last_valid_index()]
                E6_TR=REBAL_DATA['E6'].loc[REBAL_DATA['E6'].last_valid_index()]
                E7_TR=REBAL_DATA['E7'].loc[REBAL_DATA['E7'].last_valid_index()]
                E8_TR=REBAL_DATA['E8'].loc[REBAL_DATA['E8'].last_valid_index()]
        
            
               
                    
                if P[-1]<(sum(P[-3:])/3) and Q[-1]>(sum(Q[-6:])/6) and E1_TR==0:
                    SIGNAL_1_TRADE=1
                elif P[-1]>(sum(P[-3:])/3) and Q[-1]<(sum(Q[-6:])/6) and E1_TR>0:
                    SIGNAL_1_TRADE=1
                else:
                    SIGNAL_1_TRADE=0
                
            
                if PRICE_OSC<0 and BREADTH > 0 and E2_TR==0:
                    SIGNAL_2_TRADE=1
                elif PRICE_OSC>0 and BREADTH < 0 and E2_TR>0:
                    SIGNAL_2_TRADE=1
                elif X_PRICE_OSC>0 and X_BREADTH < 0 and E2_TR>0:
                    SIGNAL_2_TRADE=1
                else:
                    SIGNAL_2_TRADE=0
                
                SIGNAL=SIGNAL_1_TRADE+SIGNAL_2_TRADE+SIGNAL_3_LONG+SIGNAL_4_LONG+SIGNAL_5_LONG+SIGNAL_6_LONG
                
                if SIGNAL_1_TRADE==1:
                    Strategy_1_rebal=True
                else:
                    Strategy_1_rebal=False
            
                if SIGNAL_2_TRADE==1:
                    Strategy_2_rebal=True
                else:
                    Strategy_2_rebal=False
                
            if SIGNAL>0:
                Index_rebal=Calc_Date
            else:
                Index_rebal=False
            
            if SIGNAL>0:
                if P[-1]<(sum(P[-3:])/3) and Q[-1]>(sum(Q[-6:])/6):
                    E1=0.25
                elif P[-1]>(sum(P[-3:])/3) and Q[-1]<(sum(Q[-6:])/6):
                    E1=0
                else:
                    if Calc_Date==Start_Date:
                        E1=0
                    else:
                        E1=E1_TR
        
                E7=0.25-E1
                
                if PRICE_OSC<0 and BREADTH > 0:
                    E2=0.25
                elif PRICE_OSC>0 and BREADTH < 0:
                    E2=0
                elif X_PRICE_OSC>0 and X_BREADTH < 0:
                    E2=0
                else:
                    if Calc_Date==Start_Date:
                        E2=0
                    else:
                        E2=E2_TR
                
                E8=0.25-E2
                
                
            if Calc_Date==Start_Date:
                if P[-1]<(sum(P[-3:])/3) and Q[-1]>(sum(Q[-6:])/6) and E1==0:
                    SIGNAL_1_TRADE=1
                elif P[-1]>(sum(P[-3:])/3) and Q[-1]<(sum(Q[-6:])/6) and E1>0:
                    SIGNAL_1_TRADE=1
                else:
                    SIGNAL_1_TRADE=0
                
          
               
            
                if PRICE_OSC<0 and BREADTH > 0 and E2==0:
                    SIGNAL_2_TRADE=1
                elif PRICE_OSC>0 and BREADTH < 0 and E2>0:
                    SIGNAL_2_TRADE=1
                elif X_PRICE_OSC>0 and X_BREADTH < 0 and E2>0:
                    SIGNAL_2_TRADE=1
                else:
                    SIGNAL_2_TRADE=0
                    
                    
            
            if Index_rebal==Strategy_3_rebal:
                if SIGNAL_3_LONG==1:
                    E3=0.25
                else:
                    E3=0
                if SIGNAL_4_LONG==1:
                    E4=0.25
                else:
                    E4=0
                if SIGNAL_5_LONG==1:
                    E5=0.25
                else:
                    E5=0
                if SIGNAL_6_LONG==1:
                    E6=0.25
                else:
                    E6=0
            else:
                E3=E3_TR
                E4=E4_TR
                E5=E5_TR
                E6=E6_TR
            
            if Calc_Date==Strategy_3_rebal:
                Strategy_3_rebal=True
            elif Calc_Date==Start_Date:
                Strategy_3_rebal=True
            else:
                Strategy_3_rebal=False
              
            if Index_rebal==Calc_Date:
                Index_rebal=True
            
            if SIGNAL>0:
                csvData = pd.DataFrame([[Calc_Date,E1,E2,E3,E4,E5,E6,E7,E8,Calc_Date_P1,Index_rebal,Strategy_1_rebal,Strategy_2_rebal,Strategy_3_rebal,PRICE_OSC,X_PRICE_OSC,BREADTH,X_BREADTH,SHARPE_1,SHARPE_2,SHARPE_3,SHARPE_4,SIGNAL_1_TRADE,SIGNAL_2_TRADE,SIGNAL_3_LONG,SIGNAL_4_LONG,SIGNAL_5_LONG,SIGNAL_6_LONG,E1+E3,E2+E4,E5,E6,E7+E8]])
            else:
                csvData = pd.DataFrame([[Calc_Date,E1_TR,E2_TR,E3_TR,E4_TR,E5_TR,E6_TR,E7_TR,E8_TR,Calc_Date_P1,Index_rebal,Strategy_1_rebal,Strategy_2_rebal,Strategy_3_rebal,PRICE_OSC,X_PRICE_OSC,BREADTH,X_BREADTH,SHARPE_1,SHARPE_2,SHARPE_3,SHARPE_4,SIGNAL_1_TRADE,SIGNAL_2_TRADE,SIGNAL_3_LONG,SIGNAL_4_LONG,SIGNAL_5_LONG,SIGNAL_6_LONG,E1_TR+E3_TR,E2_TR+E4_TR,E5_TR,E6_TR,E7_TR+E8_TR]])
            
            # Exporting measurse in excel sheet
            if Calc_Date==Start_Date:
                append_df_to_excel("C:\\TUTTLE_AOL\\Exposure_3.xlsx",csvData , sheet_name='Measures', header=['Calc_Date','E1','E2','E3','E4','E5','E6','E7','E8','Effective_Date','Index_Rebal','Strategy_1_rebal','Strategy_2_rebal','Strategy_3_rebal','PRICE_OSC','X_PRICE_OSC','BREADTH','X_BREADTH','SHARPE_1','SHARPE_2','SHARPE_3','SHARPE_4','SIGNAL_1_TRADE','SIGNAL_2_TRADE','SIGNAL_3_LONG','SIGNAL_4_LONG','SIGNAL_5_LONG','SIGNAL_6_LONG','TLT_EXPO','SPY_EXPO','QQQ_EXPO','IWM_EXPO','SHY_EXPO'], index=False)
            else:
                append_df_to_excel("C:\\TUTTLE_AOL\\Exposure_3.xlsx",csvData , sheet_name='Measures', header=None, index=False)
            
             # Exporting Prices in excel sheet
            if Calc_Date==Start_Date:
                df_price1=TLT_DATA.iloc[::-1]
                df_price1=df_price1.tail(66)
                df_price2=SPY_DATA.iloc[::-1]
                df_price2=df_price2.tail(66)
                df_price3=QQQ_DATA.iloc[::-1]
                df_price3=df_price3.tail(66)
                df_price4=IWM_DATA.iloc[::-1]
                df_price4=df_price4.tail(66)
                df_price5=XLU_DATA.iloc[::-1]
                df_price5=df_price5.tail(6)
                df_3=df_3.rename(columns = {"DATE": "date"})
                def date_con(datetim):
                    return pd.Timestamp(datetim)
                df_3['date']=df_3['date'].apply(date_con)
                
                def date_convert(datetim):
                    return datetim.date()
                df_3['date']=df_3['date'].apply(date_convert)
                df_price6=df_3.tail(28)
                df_11=df_price1.loc[:,df_price1.columns.isin(['date'])]
                
                
                
                
                Price=pd.merge(df_11,df_price1.loc[:,df_price1.columns.isin(['date','indxx_value'])]).merge(df_price2.loc[:,df_price2.columns.isin(['date','indxx_value'])],how='left',on=['date']).merge(df_price3.loc[:,df_price3.columns.isin(['date','indxx_value'])],how='left',on=['date']).merge(df_price4.loc[:,df_price4.columns.isin(['date','indxx_value'])],how='left',on=['date']).merge(df_price5.loc[:,df_price5.columns.isin(['date','indxx_value'])],how='left',on=['date']).merge(df_price6.loc[:,df_price6.columns.isin(['date','Value','Value1'])],how='left',on=['date'])
                
                append_df_to_excel("C:\\TUTTLE_AOL\\Exposure_3.xlsx",Price , sheet_name='Prices', header=['Date','TLT Adjusted Price','SPY Adjusted Price','QQQ Adjusted Price','IWM Adjusted Price','XLU Adjusted Price','Decline_Issues','Advance_Issues'], index=False)
                
            else:
                df_price1=TLT_DATA.head(1)
                df_price2=SPY_DATA.head(1)
                df_price3=QQQ_DATA.head(1)
                df_price4=IWM_DATA.head(1)
                df_price5=XLU_DATA.head(1)
                df_3=df_3.rename(columns = {"DATE": "date"})
                def date_con(datetim):
                    return pd.Timestamp(datetim)
                df_3['date']=df_3['date'].apply(date_con)
                
                def date_convert(datetim):
                    return datetim.date()
                df_3['date']=df_3['date'].apply(date_convert)
                
                df_price6=df_3.tail(1)
                df_11=df_price1.loc[:,df_price1.columns.isin(['date'])]
                
                
                
                
                Price=pd.merge(df_11,df_price1.loc[:,df_price1.columns.isin(['date','indxx_value'])]).merge(df_price2.loc[:,df_price2.columns.isin(['date','indxx_value'])],how='left',on=['date']).merge(df_price3.loc[:,df_price3.columns.isin(['date','indxx_value'])],how='left',on=['date']).merge(df_price4.loc[:,df_price4.columns.isin(['date','indxx_value'])],how='left',on=['date']).merge(df_price5.loc[:,df_price5.columns.isin(['date','indxx_value'])],how='left',on=['date']).merge(df_price6.loc[:,df_price6.columns.isin(['date','Value','Value1'])],how='left',on=['date'])
                append_df_to_excel("C:\\TUTTLE_AOL\\Exposure_3.xlsx",Price , sheet_name='Prices', header=None, index=False)
                
            mycursor.execute('''select * from admin_icai5.tbl_prices_local_curr a where a.ticker in ("TLT US Equity","SPY US Equity","QQQ US Equity","IWM US Equity","SHY US Equity") and date=%s ''' ,(Calc_Date,))    
            myresult = mycursor.fetchall()
            column=column = [d[0] for d in mycursor.description]
            df = pd.DataFrame(myresult,columns=column)
            df_1=pd.DataFrame([[Calc_Date,float(df.query("ticker=='TLT US Equity'")['price'].values[0]),float(df.query("ticker=='SPY US Equity'")['price'].values[0]),float(df.query("ticker=='QQQ US Equity'")['price'].values[0]),float(df.query("ticker=='IWM US Equity'")['price'].values[0]),float(df.query("ticker=='SHY US Equity'")['price'].values[0])]])
    
            Start_Date_1 = dt.date(year = 2019, month = 7, day = 9)
            if Calc_Date == Start_Date_1:
                append_df_to_excel("C:\\TUTTLE_AOL\\Exposure_3.xlsx",df_1 , sheet_name='Raw_Prices', header=['DATE','TLT_PRICE','SPY_PRICE','QQQ_PRICE','IWM_PRICE','SHY_PRICE'], index=False)
            else:
                append_df_to_excel("C:\\TUTTLE_AOL\\Exposure_3.xlsx",df_1 , sheet_name='Raw_Prices', header=None, index=False)
        
        
        
        #updating shares on rebalancing dates
        Start_Date_1 = dt.date(year = 2019, month = 7, day = 9)
        if Index_rebal==True: 
            mydb = mysql.connector.connect(
              host="146.20.65.208",
              user="admin5",
              passwd="Admin@1234",
              database="admin_icai5"
            )
            
            mycursor = mydb.cursor()
            
         
            
            if Calc_Date == Start_Date_1:
                index_value =  float(df_index_value['indxx_value'].values[0])
            else:
                mycursor.execute('''SELECT * FROM admin_icai5.tbl_indxx_value a where a.code in ("TUTTLE")   and a.date =%s ''',(Calc_Date,))
                myresult = mycursor.fetchall()
                column=column = [d[0] for d in mycursor.description]
                df_index_value = pd.DataFrame(myresult,columns=column)
                index_value = float(df_index_value['indxx_value'].values[0])
                divisor = float(df_index_value['newdivisor'].values[0])
            
            mycursor.execute('''select * from admin_icai5.tbl_prices_local_curr a where a.ticker in ("TLT US Equity","SPY US Equity","QQQ US Equity","IWM US Equity","SHY US Equity") and date=%s ''' ,(Calc_Date,))    
            myresult = mycursor.fetchall()
            column=column = [d[0] for d in mycursor.description]
            df = pd.DataFrame(myresult,columns=column)
            
            TLT_Share = index_value*(E1+E3)*divisor/float(df.query("ticker=='TLT US Equity'")['price'].values[0])
            SPY_Share = index_value*(E2+E4)*divisor/float(df.query("ticker=='SPY US Equity'")['price'].values[0])
            QQQ_Share = index_value*(E5)*divisor/float(df.query("ticker=='QQQ US Equity'")['price'].values[0])
            IWM_Share = index_value*(E6)*divisor/float(df.query("ticker=='IWM US Equity'")['price'].values[0])
            SHY_Share = index_value*(E7+E8)*divisor/float(df.query("ticker=='SHY US Equity'")['price'].values[0])
            
            df_1=pd.DataFrame([[Calc_Date,TLT_Share,SPY_Share,QQQ_Share,IWM_Share,SHY_Share,float(df.query("ticker=='TLT US Equity'")['price'].values[0]),float(df.query("ticker=='SPY US Equity'")['price'].values[0]),float(df.query("ticker=='QQQ US Equity'")['price'].values[0]),float(df.query("ticker=='IWM US Equity'")['price'].values[0]),float(df.query("ticker=='SHY US Equity'")['price'].values[0])]])
            
        
            mycursor.execute('''update admin_icai5.tbl_share set share=%s, date=%s where indxx_id=560 and isin in ("US4642874329")   ''',(TLT_Share,Calc_Date,))
            mycursor.execute('''update admin_icai5.tbl_share set share=%s,  date=%s where indxx_id=560 and isin in ("US78462F1030")   ''',(SPY_Share,Calc_Date,))
            mycursor.execute('''update admin_icai5.tbl_share set share=%s, date=%s where  indxx_id=560 and isin in ("US46090E1038")   ''',(QQQ_Share,Calc_Date,))
            mycursor.execute('''update admin_icai5.tbl_share set share=%s,  date=%s where  indxx_id=560 and isin ="US4642876555"   ''',(IWM_Share,Calc_Date,))
            mycursor.execute('''update admin_icai5.tbl_share set share=%s , date=%s where indxx_id=560 and isin ="US4642874576"   ''',(SHY_Share,Calc_Date,))
            mydb.commit()
            
            if Calc_Date == Start_Date_1:
                append_df_to_excel("C:\\TUTTLE_AOL\\Exposure_3.xlsx",df_1 , sheet_name='Rebalancing Shares&Prices', header=['Date','TLT_SHARE','SPY_SHARE','QQQ_SHARE','IWM_Share','SHY_Share','TLT_PRICE','SPY_PRICE','QQQ_PRICE','IWM_PRICE','SHY_PRICE'], index=False)
            else:
                append_df_to_excel("C:\\TUTTLE_AOL\\Exposure_3.xlsx",df_1 , sheet_name='Rebalancing Shares&Prices', header=None, index=False)
    
        body = 'Subject: Tuttle Response File'+'\n' + '\nHello,\n \nThe code has been executed successfully\n' + '\nHave a nice day!'
        send_mail(body)
    except Exception as error:
        body = 'Subject: Tuttle Response File'+'\n' + '\nHello, \n\n'+'Error is'+ str(error) + '\nHave a nice day!'
        send_mail(body)

   
######## see untitled_5 of jupyter
        